# -*- coding: utf-8 -*-
# @Time    : 2022/6/6 14:22
# @Author  : 臧成龙
# @FileName: usual.py
# @Software: PyCharm
# -*- coding: utf-8 -*-
import os
from datetime import datetime
from urllib.parse import unquote

import openpyxl
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from fuadmin.settings import BASE_DIR, STATIC_URL
from ninja import Schema
from openpyxl import load_workbook

from .fu_auth import data_permission
from .fu_ninja import FuFilters
from .fu_response import FuResponse
from .usual import get_user_info_from_token


class ImportSchema(Schema):
    path: str


def create(request, data, model):
    """
    创建新记录的函数。

    参数:
    - request: 请求对象，用于获取用户信息。
    - data: 创建记录的数据，可以是字典或者支持dict()方法的对象。
    - model: Django ORM模型，指定要创建的对象类型。

    返回值:
    - 创建成功的对象查询集合。
    """
    if not isinstance(data, dict):
        # 如果data不是字典类型，则转换为字典
        data = data.dict()
    user_info = get_user_info_from_token(request)
    # 从请求中提取用户信息，并添加到数据中作为创建人、修改者和所属部门信息
    data['creator_id'] = user_info['id']
    data['modifier'] = user_info['name']
    data['belong_dept'] = user_info['dept']
    # 使用提供的模型和数据创建新记录
    query_set = model.objects.create(**data)
    return query_set


def batch_create(request, data, model):
    """
    批量创建模型实例。

    参数:
    - request: HTTP请求对象，用于获取用户信息。
    - data: 一个包含创建数据的列表，每个元素可以是字典或者具有dict方法的对象。
    - model: Django模型类，用于实例化和批量创建。

    返回值:
    - query_set: 批量创建后的模型实例查询集。
    """
    user_info = get_user_info_from_token(request)  # 从请求中获取用户信息
    data_list = []
    for item in data:
        if not isinstance(item, dict):
            item = item.dict()  # 如果item不是字典，则转换为字典

        # 为每个创建的数据项添加默认的创建人、修改者和所属部门信息
        item['creator_id'] = user_info['id']
        item['modifier'] = user_info['name']
        item['belong_dept'] = user_info['dept']
        data_list.append(model(**item))  # 根据字典内容实例化模型对象并添加到列表中
    query_set = model.objects.bulk_create(data_list)  # 批量创建模型实例
    return query_set


def delete(id, model):
    """
    根据提供的ID和模型删除对象实例。

    参数:
    - id: 要删除的对象的ID。
    - model: 对象所属的模型类。

    返回值:
    - 无返回值。
    """
    # 根据ID和模型获取对象实例，如果找不到则返回404错误页面
    instance = get_object_or_404(model, id=id)
    # 删除对象实例
    instance.delete()
    pass


def update(request, id, data, model):
    """
    更新给定模型实例的数据。

    参数:
    - request: HTTP请求对象，用于获取用户信息。
    - id: 要更新的模型实例的ID。
    - data: 包含更新数据的实例，应能转换为字典格式。
    - model: 要更新的模型类。

    返回值:
    - 更新后的模型实例。
    """
    dict_data = data.dict()  # 将data转换为字典格式
    user_info = get_user_info_from_token(request)  # 从请求中获取用户信息
    # 为更新的数据添加修改者信息
    dict_data['modifier'] = user_info['name']
    instance = get_object_or_404(model, id=id)  # 获取指定ID的模型实例
    # 遍历字典，将更新的数据设置到模型实例上
    for attr, value in dict_data.items():
        setattr(instance, attr, value)
    instance.save()  # 保存更新
    return instance  # 返回更新后的实例


def retrieve(request, model, filters: FuFilters = FuFilters()):
    """
    根据提供的过滤条件从数据库中检索模型实例。

    参数:
    - request: HttpRequest对象，用于获取请求信息。
    - model: Django模型类，指定要检索的数据模型。
    - filters: FuFilters类的实例，包含过滤条件。默认为FuFilters()，即无条件过滤。

    返回值:
    - query_set: 一个Django QuerySet对象，包含根据过滤条件检索到的模型实例。
    """
    # 根据请求和过滤条件应用数据权限控制
    filters = data_permission(request, filters)
    if filters is not None:
        # 将filters中的空字符串值转换为None
        for attr, value in filters.__dict__.items():
            if getattr(filters, attr) == '':
                setattr(filters, attr, None)
        # 使用过滤条件查询模型实例
        query_set = model.objects.filter(**filters.dict(exclude_none=True))
    else:
        # 如果没有有效的过滤条件，则返回所有模型实例
        query_set = model.objects.all()
    return query_set


def export_data(request, model, scheme, export_fields):
    """
    导出数据为Excel文件。

    参数:
    - request: HttpRequest对象，表示客户端请求。
    - model: Django模型类，指定要导出数据的模型。
    - scheme: 表示数据转换规则的对象，用于将ORM对象转换为字典。
    - export_fields: 包含要导出的字段名的列表。

    返回值:
    - FileResponse对象，提供下载Excel文件。
    """

    title_dict = {}
    # 根据export_fields列表获取字段的显示名称
    for field in export_fields:
        field_obj = getattr(model, field).field
        title_dict[field] = field_obj.help_text

    qs = retrieve(request, model)
    list_data = []
    # 将查询集中的每一项转换为指定格式的字典
    for qs_item in qs:
        qs_item = scheme.from_orm(qs_item)
        dict_data = {}
        for item, value in title_dict.items():
            dict_data[value] = getattr(qs_item, item)
        list_data.append(dict_data)

    # 创建并初始化Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    # 向Excel中写入数据
    for index, data in enumerate(list_data):
        if index == 0:
            # 写入表头
            ws.append(list(data.keys()))
        ws.append(list(data.values()))

    # 生成唯一的文件名并设置文件路径
    file_name = datetime.now().strftime('%Y%m%d%H%M%S%f') + '.xlsx'
    current_ymd = datetime.now().strftime('%Y%m%d')
    file_path = os.path.join(STATIC_URL, current_ymd)
    file_url = os.path.join(file_path, file_name)
    # 确保导出文件的目录存在
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # 保存Excel文件到指定路径
    wb.save(file_url)
    # 返回供下载的文件响应
    return FileResponse(open(file_url, "rb"), as_attachment=True)


def import_data(request, model, scheme, data, import_fields):
    """
    导入数据到指定模型

    参数:
    - request: HttpRequest对象，表示客户端请求
    - model: Django模型类，数据将被导入到这个模型
    - scheme: 一个函数，用于根据给定的数据字典创建模型实例
    - data: 包含要导入文件信息的对象，比如上传的Excel文件
    - import_fields: 一个列表，指定模型中需要导入的字段名

    返回值:
    - FuResponse对象，包含导入结果的消息
    """
    title_dict = {}  # 字段名与Excel列对应的字典
    for field in import_fields:
        field_obj = getattr(model, field).field
        title_dict[field_obj.help_text] = field_obj.column
    # 文件路径处理
    file_path = str(BASE_DIR) + unquote(data.path)
    # 加载Excel工作簿
    wb = load_workbook(file_path)
    ws = wb.active  # 获取活动工作表
    title_value = []
    for index_row, row in enumerate(ws.values):
        if index_row == 0:
            title_value = row  # 读取Excel表头
        else:
            dict_data = {}  # 存储每一行数据转换后的字典
            for index, cell in enumerate(row):
                title_cell = title_value[index]
                value = title_dict.get(title_cell)  # 根据表头查找对应字段
                if value is not None:
                    dict_data[value] = cell
            print(dict_data)  # 打印处理后的数据，用于调试
            data = scheme(**dict_data)  # 根据处理后的字典创建模型实例
            create(request, data, model)  # 在数据库中创建模型实例
    return FuResponse(msg='导入成功')  # 返回成功消息
